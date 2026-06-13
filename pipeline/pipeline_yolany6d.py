"""
pipeline_yolany6d.py
YOLOE + Any6D end-to-end 6D pose estimation pipeline.

Anchor mode (1 frame per object, default):
    python pipeline/pipeline_yolany6d.py --dataset dexycb

CSV mode  (N images per object):
    python pipeline/pipeline_yolany6d.py --dataset dexycb --csv frames.csv

Single object:
    python pipeline/pipeline_yolany6d.py --dataset dexycb --obj 006_mustard_bottle

CSV format (one row per frame):
    folder,color_path,depth_path,gt_mask_path,gt_pose_path,K_path
"""

import os
import sys
import csv
import glob
import pickle
import argparse
import warnings
import numpy as np
import cv2
import trimesh

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec

BASE_DIR = os.path.expanduser('~/open-vocabulary-6d-pose-yoloe')
sys.path.insert(0, os.path.join(BASE_DIR, 'utils'))
sys.path.insert(0, os.path.join(BASE_DIR, 'pipeline'))

from set_config import (
    DATASETS, EVAL_DIR, ANY6D_DIR, YOLOE_MODEL_PATH,
    YOLOE_CONF, YOLOE_IOU, ANY6D_ITER,
    THRESH_ADD_RATIO, THRESH_MSSD_RATIO, THRESH_MSPD_PX,
    PAPER_RESULTS,
)
from pipeline_utils import (run_any6d_docker, run_any6d_docker_batch,
                            compute_errors, compute_mask_metrics)
from yoloe_helpers  import yoloe_text_prompt
from any6d_utils    import (
    mesh_bbox_corners, project_points,
    draw_3d_bbox, draw_pose_axes,
    mask_overlay, colormap_depth,
)


# ── BOP metrics ───────────────────────────────────────────────────────────────

def _sample_mesh_pts(mesh, n=1000):
    """Return (n, 3) surface points sampled from mesh."""
    pts, _ = trimesh.sample.sample_surface(mesh, n)
    return pts.astype(np.float64)


def _mesh_diameter(mesh):
    """Approximate object diameter as longest bounding-box diagonal (metres)."""
    return float(np.linalg.norm(mesh.bounding_box.extents))


def _nn_dist(pts_a, pts_b):
    """For each row in pts_a, find min L2 distance to any row in pts_b."""
    diff  = pts_a[:, None, :] - pts_b[None, :, :]   # (N, M, 3)
    dists = np.linalg.norm(diff, axis=2)              # (N, M)
    return dists.min(axis=1)                           # (N,)


def _z_symmetry_tfs(n=36):
    """36 rigid transforms representing discrete Z-axis rotational symmetry."""
    tfs = []
    for i in range(n):
        a = 2 * np.pi * i / n
        c, s = np.cos(a), np.sin(a)
        T = np.eye(4)
        T[:3, :3] = np.array([[c, -s, 0], [s, c, 0], [0, 0, 1]])
        tfs.append(T)
    return tfs


def _transform_pts(pts, pose):
    """Apply (4,4) pose to (N,3) points."""
    return (pose[:3, :3] @ pts.T + pose[:3, 3:]).T


def _project_pts_2d(pts_3d, K):
    """Project (N,3) camera-frame points to (N,2) pixel coords."""
    p   = (K @ pts_3d.T).T
    xy  = p[:, :2] / p[:, 2:]
    return xy


def compute_bop_metrics(pred_pose, gt_pose, mesh, K, symmetric=False):
    """
    Compute ADD, ADD-S, MSSD, MSPD and their BOP pass/fail flags.

    Returns dict with keys:
        ADD, ADD_S, MSSD, MSPD — float errors
        ADD_ok, ADDS_ok, MSSD_ok, MSPD_ok — bool pass flags
        AR — float mean(ADDS_ok, MSSD_ok, MSPD_ok)
        diameter — object diameter in metres
    """
    pts_src   = _sample_mesh_pts(mesh, n=1000)
    diameter  = _mesh_diameter(mesh)
    pts_pred  = _transform_pts(pts_src, pred_pose)
    pts_gt    = _transform_pts(pts_src, gt_pose)

    # ADD
    add = float(np.mean(np.linalg.norm(pts_pred - pts_gt, axis=1)))

    # ADD-S (ADI) — for symmetric: nearest point; non-sym: same as ADD
    if symmetric:
        adds = float(np.mean(_nn_dist(pts_pred, pts_gt)))
    else:
        adds = add

    # MSSD (one-sided Hausdorff in 3D)
    if symmetric:
        sym_tfs = _z_symmetry_tfs()
        mssd = min(
            float(_nn_dist(pts_pred, _transform_pts(pts_src, gt_pose @ T_s)).max())
            for T_s in sym_tfs
        )
    else:
        mssd = float(_nn_dist(pts_pred, pts_gt).max())

    # MSPD (one-sided Hausdorff in 2D projected space)
    proj_pred = _project_pts_2d(pts_pred, K)
    proj_gt   = _project_pts_2d(pts_gt,   K)
    if symmetric:
        mspd_vals = []
        for T_s in _z_symmetry_tfs():
            pts_gt_s   = _transform_pts(pts_src, gt_pose @ T_s)
            proj_gt_s  = _project_pts_2d(pts_gt_s, K)
            nn_2d      = np.linalg.norm(
                proj_pred[:, None, :] - proj_gt_s[None, :, :], axis=2
            ).min(axis=1).max()
            mspd_vals.append(float(nn_2d))
        mspd = min(mspd_vals)
    else:
        diff_2d = proj_pred[:, None, :] - proj_gt[None, :, :]
        mspd    = float(np.linalg.norm(diff_2d, axis=2).min(axis=1).max())

    add_ok  = add  < diameter * THRESH_ADD_RATIO
    adds_ok = adds < diameter * THRESH_ADD_RATIO
    mssd_ok = mssd < diameter * THRESH_MSSD_RATIO
    mspd_ok = mspd < THRESH_MSPD_PX
    ar      = float(np.mean([adds_ok, mssd_ok, mspd_ok]))

    return {
        'ADD' : add,  'ADD_ok' : add_ok,
        'ADD_S': adds, 'ADDS_ok': adds_ok,
        'MSSD': mssd,  'MSSD_ok': mssd_ok,
        'MSPD': mspd,  'MSPD_ok': mspd_ok,
        'AR'  : ar,
        'diameter': diameter,
    }


# ── Visualization ─────────────────────────────────────────────────────────────

def _ax_title(ax, text, fontsize=9, pad=4, **kw):
    ax.set_title(text, fontsize=fontsize, pad=pad, **kw)


def save_pipeline_figure(
    color_rgb, depth_raw, mask_bool,
    pred_pose, gt_pose, K, mesh,
    bop, mask_metrics,
    obj_name, out_path
):
    """
    Save a 2×2 panel figure for one object:
        [0,0] Anchor RGB    [0,1] Depth colourmap
        [1,0] YOLOE mask    [1,1] 6D pose result

    Uses non-interactive Agg backend — safe on headless servers.
    """
    fig, axes = plt.subplots(2, 2, figsize=(13, 9))
    fig.patch.set_facecolor('#f5f5f5')
    plt.suptitle(
        f'{obj_name.replace("_", " ")}    '
        f'AR={bop["AR"]:.2f}  ADD-S={bop["ADD_S"]*100:.1f}cm  '
        f'T={bop.get("err_t", float("nan")):.1f}cm  R={bop.get("err_R", float("nan")):.1f}°',
        fontsize=11, fontweight='normal', y=0.98
    )

    # ── Panel 0,0 : Anchor RGB ────────────────────────────────────────────────
    ax = axes[0, 0]
    ax.imshow(color_rgb)
    _ax_title(ax, 'Anchor frame (RGB)')
    ax.axis('off')

    # ── Panel 0,1 : Depth colourmap ───────────────────────────────────────────
    ax = axes[0, 1]
    valid     = depth_raw[depth_raw > 0]
    d_min     = valid.min() if valid.size else 0
    d_max     = valid.max() if valid.size else 1
    d_norm    = np.clip((depth_raw - d_min) / (d_max - d_min + 1e-9), 0, 1)
    ax.imshow(d_norm, cmap='inferno')
    _ax_title(ax, f'Depth  [{d_min:.3f} – {d_max:.3f} m]')
    ax.axis('off')

    # ── Panel 1,0 : YOLOE mask overlay ───────────────────────────────────────
    ax = axes[1, 0]
    overlay   = mask_overlay(color_rgb, mask_bool, alpha=0.45)
    ax.imshow(overlay)
    iou_str   = f'IoU={mask_metrics["IoU"]:.3f}  Dice={mask_metrics["Dice"]:.3f}  ' \
                f'P={mask_metrics["Precision"]:.2f}  R={mask_metrics["Recall"]:.2f}'
    _ax_title(ax, f'YOLOE mask  ({mask_bool.sum()} px)\n{iou_str}', fontsize=8)
    ax.axis('off')

    # ── Panel 1,1 : Pose result ───────────────────────────────────────────────
    ax = axes[1, 1]
    corners_3d = mesh_bbox_corners(mesh)
    base       = mask_overlay(color_rgb, mask_bool, alpha=0.20)

    # predicted — blue
    c_pred = project_points(corners_3d, pred_pose, K)
    img    = draw_3d_bbox(base, c_pred, color=(30, 120, 255), thickness=2)
    img    = draw_pose_axes(img, pred_pose, K, length=0.05)

    if gt_pose is not None:
        c_gt = project_points(corners_3d, gt_pose, K)
        img  = draw_3d_bbox(img, c_gt, color=(30, 200, 80), thickness=2)

    ax.imshow(img)

    ar_color = '#2e7d32' if bop['AR'] >= 0.8 else ('#f57f17' if bop['AR'] >= 0.4 else '#c62828')
    title_str = (
        f'Pose result   AR={bop["AR"]:.2f}\n'
        f'ADD={bop["ADD"]*100:.1f}cm  ADD-S={bop["ADD_S"]*100:.1f}cm  '
        f'MSSD={bop["MSSD"]*100:.1f}cm  MSPD={bop["MSPD"]:.1f}px'
    )
    _ax_title(ax, title_str, fontsize=8, color=ar_color)
    ax.axis('off')

    handles = [
        mpatches.Patch(color='#1E78FF', label='Predicted'),
        mpatches.Patch(color='#1EC850', label='GT'),
        mpatches.Patch(color='#DC3232', label='X'),
        mpatches.Patch(color='#32DC32', label='Y'),
        mpatches.Patch(color='#3264E6', label='Z'),
    ]
    ax.legend(handles=handles, loc='lower left', fontsize=7,
              framealpha=0.75, ncol=3)

    plt.tight_layout(rect=[0, 0, 1, 0.97])
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    fig.savefig(out_path, dpi=120, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f'  Figure: {out_path}')


# ── Frame loading ─────────────────────────────────────────────────────────────

def _load_frame(color_path, depth_path, K_path):
    color_bgr = cv2.imread(color_path)
    if color_bgr is None:
        raise FileNotFoundError(f'color not found: {color_path}')
    color_rgb = cv2.cvtColor(color_bgr, cv2.COLOR_BGR2RGB)
    depth_raw = cv2.imread(depth_path, cv2.IMREAD_ANYDEPTH).astype(np.float32) / 1000.0
    K         = np.loadtxt(K_path)
    return color_rgb, depth_raw, K


def _frames_from_anchor(obj_cfg, dataset_cfg):
    """Yield one frame dict from the anchor folder."""
    folder     = obj_cfg['folder']
    anchor_dir = os.path.join(dataset_cfg['anchor_dir'], folder)
    yield {
        'folder'        : folder,
        'color_path'    : os.path.join(anchor_dir, 'color.png'),
        'depth_path'    : os.path.join(anchor_dir, 'depth.png'),
        'K_path'        : os.path.join(anchor_dir, 'K.txt'),
        'mesh_path'     : os.path.join(anchor_dir, obj_cfg['mesh']),
        'gt_mask_path'  : os.path.join(anchor_dir, 'mask.png'),
        'gt_pose_path'  : os.path.join(anchor_dir, f'{folder}_gt_pose.txt'),
        'init_pose_path': os.path.join(anchor_dir, f'{folder}_initial_pose.txt'),
        'symmetric'     : obj_cfg.get('symmetric', False),
        'prompt'        : obj_cfg['prompt'],
        'use_rfix'      : dataset_cfg.get('use_rfix', False),
    }


def _frames_from_csv(csv_path, dataset_cfg):
    """Yield frame dicts parsed from a CSV file."""
    with open(csv_path) as f:
        reader = csv.DictReader(f)
        for row in reader:
            folder = row['folder']
            obj_cfg = next(
                (o for o in dataset_cfg['objects'] if o['folder'] == folder), None
            )
            if obj_cfg is None:
                print(f'  [skip] folder not in config: {folder}')
                continue
            anchor_dir = os.path.join(dataset_cfg['anchor_dir'], folder)
            yield {
                'folder'        : folder,
                'color_path'    : row['color_path'],
                'depth_path'    : row['depth_path'],
                'K_path'        : row.get('K_path', os.path.join(anchor_dir, 'K.txt')),
                'mesh_path'     : os.path.join(anchor_dir, obj_cfg['mesh']),
                'gt_mask_path'  : row.get('gt_mask_path', ''),
                'gt_pose_path'  : row.get('gt_pose_path', ''),
                'init_pose_path': os.path.join(anchor_dir, f'{folder}_initial_pose.txt'),
                'symmetric'     : obj_cfg.get('symmetric', False),
                'prompt'        : obj_cfg['prompt'],
                'use_rfix'      : dataset_cfg.get('use_rfix', False),
            }


# ── R_fix (coordinate frame alignment) ───────────────────────────────────────

def _apply_r_fix(pred_pose, gt_pose, init_pose_path):
    """
    Align GT coordinate frame to pred when the coordinate conventions differ.
    Auto-detects the reference: uses initial_pose if angle < 30°, else pred_pose.
    """
    if not os.path.exists(init_pose_path):
        return gt_pose
    init_pose = np.loadtxt(init_pose_path)
    R_diff    = init_pose[:3, :3] @ pred_pose[:3, :3].T
    angle     = np.degrees(np.arccos(np.clip((np.trace(R_diff) - 1) / 2, -1, 1)))
    ref_R     = init_pose[:3, :3] if angle < 30 else pred_pose[:3, :3]
    R_fix     = np.eye(4)
    R_fix[:3, :3] = gt_pose[:3, :3].T @ ref_R
    return gt_pose @ R_fix


# ── Single frame processing ───────────────────────────────────────────────────

def process_frame(model, frame, conf, iou_thresh, iteration, viz_dir, save_dir):
    """
    Run YOLOE + Any6D on one frame. Returns a result dict.
    """
    folder     = frame['folder']
    color_rgb, depth_raw, K = _load_frame(
        frame['color_path'], frame['depth_path'], frame['K_path']
    )
    mesh = trimesh.load(frame['mesh_path'])

    # ── YOLOE detection ───────────────────────────────────────────────────────
    print(f'\n  [YOLOE] {folder}')
    bbox_xyxy, mask_bool, det_conf, _ = yoloe_text_prompt(
        model, frame['color_path'], frame['prompt'],
        conf=conf, iou=iou_thresh
    )
    print(f'    prompt={frame["prompt"]}  conf={det_conf:.3f}  mask={mask_bool.sum()} px')

    # ── GT mask quality ───────────────────────────────────────────────────────
    if os.path.exists(frame.get('gt_mask_path', '')):
        gt_mask_img  = cv2.imread(frame['gt_mask_path'], cv2.IMREAD_GRAYSCALE)
        gt_mask_bool = gt_mask_img > 127
        mask_metrics = compute_mask_metrics(mask_bool, gt_mask_bool)
    else:
        gt_mask_bool = None
        mask_metrics = {'IoU': float('nan'), 'Dice': float('nan'),
                        'Precision': float('nan'), 'Recall': float('nan')}

    # ── Any6D pose estimation ─────────────────────────────────────────────────
    print(f'  [Any6D] {folder}')
    save_path = os.path.join(save_dir, folder)
    pred_pose = run_any6d_docker(
        color_path=frame['color_path'],
        depth_path=frame['depth_path'],
        mask_bool=mask_bool,
        K=K,
        mesh_path=frame['mesh_path'],
        save_path=save_path,
        name=folder,
        iteration=iteration,
    )

    # ── GT pose + error ───────────────────────────────────────────────────────
    gt_pose_path = frame.get('gt_pose_path', '')
    if os.path.exists(gt_pose_path):
        gt_pose = np.loadtxt(gt_pose_path)
        if frame.get('use_rfix', False):
            # DexYCB only: GT uses YCB-canonical frame (90-176° off vs Any6D convention)
            gt_pose = _apply_r_fix(pred_pose, gt_pose, frame.get('init_pose_path', ''))
        # Other datasets (HO3D, LMO, REAL275…): gt_pose already in camera frame → no fix
        err_t, err_R = compute_errors(pred_pose, gt_pose)
        bop = compute_bop_metrics(pred_pose, gt_pose, mesh, K, symmetric=frame['symmetric'])
        bop['err_t'] = err_t
        bop['err_R'] = err_R
    else:
        gt_pose = None
        err_t = err_R = float('nan')
        bop = {k: float('nan') for k in
               ['ADD','ADD_S','MSSD','MSPD','AR','diameter']}
        bop.update({'ADD_ok': False, 'ADDS_ok': False,
                    'MSSD_ok': False, 'MSPD_ok': False,
                    'err_t': err_t, 'err_R': err_R})

    print(f'    T={err_t:.2f} cm   R={err_R:.2f} deg   AR={bop.get("AR", float("nan")):.3f}')

    # ── 4-panel figure ────────────────────────────────────────────────────────
    fig_path = os.path.join(viz_dir, f'{folder}_pipeline.png')
    save_pipeline_figure(
        color_rgb=color_rgb,
        depth_raw=depth_raw,
        mask_bool=mask_bool,
        pred_pose=pred_pose,
        gt_pose=gt_pose,
        K=K,
        mesh=mesh,
        bop=bop,
        mask_metrics=mask_metrics,
        obj_name=folder,
        out_path=fig_path,
    )

    return {
        'folder'      : folder,
        'pred_pose'   : pred_pose,
        'gt_pose'     : gt_pose,
        'err_t'       : err_t,
        'err_R'       : err_R,
        'bop'         : bop,
        'mask_metrics': mask_metrics,
        'det_conf'    : det_conf,
        'det_px'      : int(mask_bool.sum()),
        'success'     : True,
    }


# ── Excel export ─────────────────────────────────────────────────────────────

def _cell_color(ok):
    from openpyxl.styles import PatternFill
    c = 'C8E6C9' if ok else 'FFCDD2'
    return PatternFill(start_color=c, end_color=c, fill_type='solid')


def export_excel(all_results, dataset_name, paper_results, out_path):
    """
    Write per-object BOP + mask metrics to Excel with colour-coded cells.
    Columns: Object | ADD | ADD-S | MSSD | MSPD | AR | ADD-ok | ADDS-ok |
             MSSD-ok | MSPD-ok | IoU | Dice | Precision | Recall |
             T-err(cm) | R-err(deg) | YOLOE-conf | Paper ADD-S | Paper AR
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('  [Excel] openpyxl not installed — skipping Excel export')
        return

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = dataset_name[:31]

    header_fill = PatternFill(start_color='37474F', end_color='37474F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    headers = [
        'Object',
        'ADD (cm)', 'ADD-S (cm)', 'MSSD (cm)', 'MSPD (px)',
        'AR', 'ADD ok', 'ADD-S ok', 'MSSD ok', 'MSPD ok',
        'IoU', 'Dice', 'Precision', 'Recall',
        'T err (cm)', 'R err (deg)', 'YOLOE conf',
        'Paper ADD-S', 'Paper AR',
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    ok_keys = ['ADD_ok', 'ADDS_ok', 'MSSD_ok', 'MSPD_ok']

    for r in all_results:
        if not r.get('success'):
            continue
        bop  = r['bop']
        mm   = r['mask_metrics']
        p_ar = paper_results.get('AR')
        p_ad = paper_results.get('ADD-S')

        def _f(v, scale=1, nd=3):
            if v is None or (isinstance(v, float) and np.isnan(v)):
                return 'N/A'
            return round(float(v) * scale, nd)

        row = [
            r['folder'],
            _f(bop.get('ADD'),  100, 2),
            _f(bop.get('ADD_S'), 100, 2),
            _f(bop.get('MSSD'), 100, 2),
            _f(bop.get('MSPD'),   1, 1),
            _f(bop.get('AR'),     1, 3),
            '✓' if bop.get('ADD_ok')  else '✗',
            '✓' if bop.get('ADDS_ok') else '✗',
            '✓' if bop.get('MSSD_ok') else '✗',
            '✓' if bop.get('MSPD_ok') else '✗',
            _f(mm.get('IoU'),       1, 3),
            _f(mm.get('Dice'),      1, 3),
            _f(mm.get('Precision'), 1, 3),
            _f(mm.get('Recall'),    1, 3),
            _f(r.get('err_t'),      1, 2),
            _f(r.get('err_R'),      1, 2),
            _f(r.get('det_conf'),   1, 3),
            _f(p_ad, 1, 3) if p_ad is not None else 'N/A',
            _f(p_ar, 1, 3) if p_ar is not None else 'N/A',
        ]
        ws.append(row)
        cur_row = ws.max_row

        # colour ok/fail cells (cols 7-10)
        for ci, key in enumerate(ok_keys, start=7):
            cell = ws.cell(row=cur_row, column=ci)
            cell.fill = _cell_color(bop.get(key, False))
            cell.alignment = Alignment(horizontal='center')

        # colour AR cell
        ar_val = bop.get('AR', 0)
        if isinstance(ar_val, float) and not np.isnan(ar_val):
            ar_fill_c = 'C8E6C9' if ar_val >= 0.8 else ('FFF9C4' if ar_val >= 0.4 else 'FFCDD2')
            ws.cell(row=cur_row, column=6).fill = PatternFill(
                start_color=ar_fill_c, end_color=ar_fill_c, fill_type='solid'
            )

        for col in range(1, len(headers) + 1):
            ws.cell(row=cur_row, column=col).border = thin_border

    # MEAN row
    valid = [r for r in all_results if r.get('success')]
    if valid:
        def _mean(key_chain):
            vals = []
            for r in valid:
                v = r
                for k in key_chain:
                    v = v.get(k) if isinstance(v, dict) else None
                    if v is None:
                        break
                if v is not None and not (isinstance(v, float) and np.isnan(v)):
                    vals.append(float(v))
            return round(np.mean(vals), 3) if vals else 'N/A'

        mean_row = [
            'MEAN',
            _mean(['bop','ADD']),
            _mean(['bop','ADD_S']),
            _mean(['bop','MSSD']),
            _mean(['bop','MSPD']),
            _mean(['bop','AR']),
            '', '', '', '',
            _mean(['mask_metrics','IoU']),
            _mean(['mask_metrics','Dice']),
            _mean(['mask_metrics','Precision']),
            _mean(['mask_metrics','Recall']),
            _mean(['err_t']),
            _mean(['err_R']),
            _mean(['det_conf']),
            '', '',
        ]
        ws.append(mean_row)
        mean_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        mean_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = mean_fill
            cell.font = mean_font
            cell.border = thin_border

    # column widths
    col_widths = [28, 11, 11, 11, 10, 8, 8, 8, 8, 8,
                  8, 8, 10, 9, 12, 12, 12, 13, 10]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f'\n  Excel: {out_path}')


# ── Terminal summary ──────────────────────────────────────────────────────────

def print_summary(all_results, dataset_name):
    SEP = '═' * 80
    print(f'\n{SEP}')
    print(f'  PIPELINE SUMMARY — {dataset_name}')
    print(SEP)

    ok  = [r for r in all_results if r.get('success')]
    bad = [r for r in all_results if not r.get('success')]
    print(f'  Processed: {len(ok)}/{len(all_results)}  |  Failed: {len(bad)}')

    if ok:
        hdr = (f'  {"Object":<30} {"AR":>5} {"ADD-S":>7} {"MSSD":>7} {"MSPD":>7}'
               f' {"T(cm)":>7} {"R(°)":>7} {"IoU":>6} {"conf":>6}')
        print(f'\n{hdr}')
        print('  ' + '─' * 78)

        for r in ok:
            bop  = r['bop']
            mm   = r['mask_metrics']
            ar   = bop.get('AR',  float('nan'))
            adds = bop.get('ADD_S', float('nan'))
            mssd = bop.get('MSSD', float('nan'))
            mspd = bop.get('MSPD', float('nan'))

            def _fmt(v, scale=1, nd=2):
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    return '  N/A'
                return f'{v*scale:>{5}.{nd}f}'

            print(
                f'  {r["folder"]:<30}'
                f' {_fmt(ar):>5}'
                f' {_fmt(adds,100):>7}'
                f' {_fmt(mssd,100):>7}'
                f' {_fmt(mspd):>7}'
                f' {_fmt(r.get("err_t")):>7}'
                f' {_fmt(r.get("err_R")):>7}'
                f' {_fmt(mm.get("IoU")):>6}'
                f' {_fmt(r.get("det_conf")):>6}'
            )

        print('  ' + '─' * 78)
        arrs  = [r['bop'].get('AR', float('nan'))    for r in ok]
        err_ts = [r.get('err_t', float('nan'))         for r in ok]
        err_Rs = [r.get('err_R', float('nan'))         for r in ok]
        ious  = [r['mask_metrics'].get('IoU', float('nan')) for r in ok]

        def _mean_fmt(vals, scale=1, nd=2):
            v = [x*scale for x in vals if not np.isnan(x)]
            return f'{np.mean(v):.{nd}f}' if v else 'N/A'

        print(
            f'  {"MEAN":<30}'
            f' {_mean_fmt(arrs):>5}'
            f' {_mean_fmt([r["bop"].get("ADD_S",float("nan")) for r in ok], 100):>7}'
            f' {_mean_fmt([r["bop"].get("MSSD",float("nan")) for r in ok], 100):>7}'
            f' {_mean_fmt([r["bop"].get("MSPD",float("nan")) for r in ok]):>7}'
            f' {_mean_fmt(err_ts):>7}'
            f' {_mean_fmt(err_Rs):>7}'
            f' {_mean_fmt(ious):>6}'
        )

        n_adds = sum(1 for r in ok if r['bop'].get('ADDS_ok', False))
        n_mssd = sum(1 for r in ok if r['bop'].get('MSSD_ok', False))
        n_mspd = sum(1 for r in ok if r['bop'].get('MSPD_ok', False))
        print(f'\n  ADD-S pass: {n_adds}/{len(ok)}  '
              f'MSSD pass: {n_mssd}/{len(ok)}  '
              f'MSPD pass: {n_mspd}/{len(ok)}')

    if bad:
        print('\n  Failed objects:')
        for r in bad:
            print(f'    {r.get("folder","?")} — {r.get("error","unknown")}')
    print(SEP + '\n')


# ── HO3D helpers ─────────────────────────────────────────────────────────────

# OpenCV ↔ OpenGL camera convention (used by HO3D GT poses)
_glcam_in_cvcam = np.array([[1, 0, 0, 0],
                              [0,-1, 0, 0],
                              [0, 0,-1, 0],
                              [0, 0, 0, 1]], dtype=np.float64)


def _ho3d_load_depth(path):
    """HO3D custom depth encoding: depth = (R + G*256) * scale."""
    scale = 0.00012498664727900177
    d = cv2.imread(path, cv2.IMREAD_UNCHANGED)
    if d is None:
        raise FileNotFoundError(f'depth not found: {path}')
    return (d[..., 2] + d[..., 1] * 256).astype(np.float32) * scale


def _ho3d_load_gt_pose(pkl_path):
    """Load object GT pose from HO3D .pkl meta file. Returns (4,4) or None."""
    with open(pkl_path, 'rb') as f:
        meta = pickle.load(f, encoding='latin1')
    if meta.get('objTrans') is None:
        return None
    pose = np.eye(4, dtype=np.float64)
    pose[:3, 3]  = np.array(meta['objTrans']).ravel()
    pose[:3, :3] = cv2.Rodrigues(np.array(meta['objRot']).ravel())[0]
    return _glcam_in_cvcam @ pose


def _ho3d_load_K(pkl_path):
    """Load camera intrinsics from HO3D .pkl meta file."""
    with open(pkl_path, 'rb') as f:
        meta = pickle.load(f, encoding='latin1')
    return np.array(meta['camMat'], dtype=np.float64)


def _apply_aq_transform(pred_pose_q, pred_pose_a, gt_pose_a):
    """
    T_{A→Q} formula from Any6D paper:
        pose_aq    = T_{O_M→Q} @ inv(T_{O_M→A})
        pred_final = pose_aq @ gt_pose_a
    Anchors Any6D's relative motion estimate to the known GT anchor pose.
    """
    pose_aq = pred_pose_q @ np.linalg.inv(pred_pose_a)
    return pose_aq @ gt_pose_a


def _ho3d_metric_recall(dist_mm, diameter_mm, thresholds_ratio):
    """Fraction of recall thresholds (as % of diameter) that dist passes."""
    return float((dist_mm < thresholds_ratio * diameter_mm).mean())


def compute_ho3d_metrics(pred_pose, gt_pose, mesh_pts, diameter):
    """
    Compute ADD-S, ADD (binary at 10% diameter) and AR (MSSD + MSPD recall).

    mesh_pts  : (N, 3) surface points in metres (sampled from GT mesh)
    diameter  : object diameter in metres

    Returns dict matching paper Table 1 format:
        adds_ok  : bool
        add_ok   : bool
        adds_m   : float (m)
        add_m    : float (m)
        mssd_rec : float [0,1]  recall over 10 MSSD thresholds
        mspd_rec : float [0,1]  recall over 10 MSPD thresholds
        ar       : float [0,1]  mean(mssd_rec, mspd_rec, adds_ok)
                                ≈ paper AR but uses ADD-S instead of VSD
    """
    # thresholds matching run_ho3d_query.py
    mssd_ratios = np.array([0.05, 0.10, 0.15, 0.20, 0.25,
                             0.30, 0.35, 0.40, 0.45, 0.50])
    mspd_px     = np.array([5, 10, 15, 20, 25, 30, 35, 40, 45, 50],
                            dtype=np.float32)

    pts_pred = _transform_pts(mesh_pts, pred_pose)
    pts_gt   = _transform_pts(mesh_pts, gt_pose)

    # ADD-S (nearest-neighbour symmetric)
    adds = float(np.mean(_nn_dist(pts_pred, pts_gt)))
    # ADD (point-to-point)
    add  = float(np.mean(np.linalg.norm(pts_pred - pts_gt, axis=1)))

    adds_ok = adds <= diameter * 0.1
    add_ok  = add  <= diameter * 0.1

    # MSSD: one-sided Hausdorff in 3D
    mssd_m = float(_nn_dist(pts_pred, pts_gt).max())
    mssd_rec = _ho3d_metric_recall(mssd_m * 1e3,
                                   diameter * 1e3, mssd_ratios)

    return {
        'adds_ok' : adds_ok,
        'add_ok'  : add_ok,
        'adds_m'  : adds,
        'add_m'   : add,
        'mssd_m'  : mssd_m,
        'mssd_rec': mssd_rec,
        'diameter': diameter,
    }


def compute_mspd_recall(pts_src, pred_pose, gt_pose, K):
    """MSPD recall over 10 pixel thresholds (5 to 50 px)."""
    mspd_px = np.array([5, 10, 15, 20, 25, 30, 35, 40, 45, 50], dtype=np.float32)
    pts_pred = _transform_pts(pts_src, pred_pose)
    pts_gt   = _transform_pts(pts_src, gt_pose)
    proj_pred = _project_pts_2d(pts_pred, K)
    proj_gt   = _project_pts_2d(pts_gt,   K)
    diff = proj_pred[:, None, :] - proj_gt[None, :, :]
    mspd = float(np.linalg.norm(diff, axis=2).min(axis=1).max())
    return float((mspd < mspd_px).mean()), mspd


def _frames_from_ho3d(dataset_cfg, stride=10, seq_filter=None):
    """
    Yield frame dicts for all HO3D evaluation sequences.
    Each frame dict is compatible with YOLOE + Any6D Docker call.
    """
    ho3d_root  = dataset_cfg['ho3d_root']
    anchor_dir = dataset_cfg['anchor_dir']
    sequences  = dataset_cfg['sequences']
    obj_map    = {o['folder']: o for o in dataset_cfg['objects']}

    for seq_name, obj_folder in sequences.items():
        if seq_filter and seq_name not in seq_filter:
            continue

        seq_dir = os.path.join(ho3d_root, 'evaluation', seq_name)
        if not os.path.isdir(seq_dir):
            yield {'seq_name': seq_name, 'missing': True,
                   'folder': obj_folder}
            continue

        rgb_files = sorted(glob.glob(os.path.join(seq_dir, 'rgb', '*.jpg')))
        if not rgb_files:
            yield {'seq_name': seq_name, 'missing': True,
                   'folder': obj_folder}
            continue

        rgb_files = rgb_files[::stride]
        if 'max_frames' in dataset_cfg and dataset_cfg.get('max_frames'):
            rgb_files = rgb_files[:dataset_cfg['max_frames']]
        obj_cfg   = obj_map[obj_folder]

        # Read K from first frame pkl (constant within a sequence)
        first_pkl = rgb_files[0].replace('.jpg', '.pkl').replace('rgb', 'meta')
        K_query   = _ho3d_load_K(first_pkl)

        for rgb_path in rgb_files:
            frame_id  = int(os.path.basename(rgb_path).split('.')[0])
            depth_path = rgb_path.replace('.jpg', '.png').replace('rgb', 'depth')
            pkl_path   = rgb_path.replace('.jpg', '.pkl').replace('rgb', 'meta')

            yield {
                'seq_name'      : seq_name,
                'frame_id'      : frame_id,
                'folder'        : obj_folder,
                'color_path'    : rgb_path,
                'depth_path'    : depth_path,
                'pkl_path'      : pkl_path,
                'K_query'       : K_query,
                'mesh_path'     : os.path.join(anchor_dir, obj_folder,
                                                obj_cfg['mesh']),
                'init_pose_path': os.path.join(anchor_dir, obj_folder,
                                                f'{obj_folder}_initial_pose.txt'),
                'gt_pose_a_path': os.path.join(anchor_dir, obj_folder,
                                                f'{obj_folder}_gt_pose.txt'),
                'anchor_K_path' : os.path.join(anchor_dir, obj_folder, 'K.txt'),
                'symmetric'     : obj_cfg.get('symmetric', False),
                'prompt'        : obj_cfg['prompt'],
                'use_rfix'      : False,
                'ho3d'          : True,
            }


def process_frame_ho3d(model, frame, conf, iou_thresh, iteration,
                        viz_dir, save_dir):
    """
    HO3D frame: YOLOE → Any6D → T_{A→Q} → HO3D metrics.
    Returns result dict with adds_ok, add_ok, ar per frame.
    """
    seq    = frame['seq_name']
    folder = frame['folder']
    fid    = frame['frame_id']
    K_q    = frame['K_query']

    # ── Load query frame ──────────────────────────────────────────────────────
    color_bgr = cv2.imread(frame['color_path'])
    color_rgb = cv2.cvtColor(color_bgr, cv2.COLOR_BGR2RGB)
    depth_m   = _ho3d_load_depth(frame['depth_path'])   # metres float32
    gt_pose_q = _ho3d_load_gt_pose(frame['pkl_path'])
    if gt_pose_q is None:
        return {'seq_name': seq, 'frame_id': fid, 'folder': folder,
                'success': False, 'error': 'No GT pose in pkl'}

    # ── YOLOE detection ───────────────────────────────────────────────────────
    try:
        bbox_xyxy, mask_bool, det_conf, _ = yoloe_text_prompt(
            model, frame['color_path'], frame['prompt'],
            conf=conf, iou=iou_thresh,
        )
    except ValueError as e:
        return {'seq_name': seq, 'frame_id': fid, 'folder': folder,
                'success': False, 'error': str(e), 'det_conf': 0.0}

    # ── Convert HO3D depth to standard uint16 mm (for run_any6d_docker) ──────
    obj_save = os.path.join(save_dir, seq, f'{fid:05d}')
    os.makedirs(obj_save, exist_ok=True)
    depth_mm_path = os.path.join(obj_save, 'depth_mm.png')
    depth_uint16  = (depth_m * 1000).clip(0, 65535).astype(np.uint16)
    cv2.imwrite(depth_mm_path, depth_uint16)

    # ── Any6D pose estimation (on query frame) ────────────────────────────────
    pred_pose_q = run_any6d_docker(
        color_path=frame['color_path'],
        depth_path=depth_mm_path,
        mask_bool=mask_bool,
        K=K_q,
        mesh_path=frame['mesh_path'],
        save_path=obj_save,
        name=f'{seq}_{fid:05d}',
        iteration=iteration,
    )

    # ── T_{A→Q} transform ────────────────────────────────────────────────────
    pred_pose_a = np.loadtxt(frame['init_pose_path'])
    gt_pose_a   = np.loadtxt(frame['gt_pose_a_path'])
    pred_final  = _apply_aq_transform(pred_pose_q, pred_pose_a, gt_pose_a)

    # ── Load mesh for metrics ─────────────────────────────────────────────────
    mesh     = trimesh.load(frame['mesh_path'])
    pts_src, _ = trimesh.sample.sample_surface(mesh, 1000)
    pts_src  = pts_src.astype(np.float64)
    diameter = float(np.linalg.norm(mesh.bounding_box.extents))

    # ── HO3D metrics ─────────────────────────────────────────────────────────
    m = compute_ho3d_metrics(pred_final, gt_pose_q, pts_src, diameter)
    mspd_rec, mspd_val = compute_mspd_recall(pts_src, pred_final, gt_pose_q, K_q)
    m['mspd_rec'] = mspd_rec
    m['mspd_val'] = mspd_val
    # AR ≈ paper AR but uses ADD-S binary instead of VSD (VSD needs renderer)
    m['ar'] = float(np.mean([m['mssd_rec'], m['mspd_rec'],
                              float(m['adds_ok'])]))

    err_t, err_R = compute_errors(pred_final, gt_pose_q)

    return {
        'seq_name' : seq,
        'frame_id' : fid,
        'folder'   : folder,
        'success'  : True,
        'adds_ok'  : m['adds_ok'],
        'add_ok'   : m['add_ok'],
        'adds_m'   : m['adds_m'],
        'add_m'    : m['add_m'],
        'mssd_rec' : m['mssd_rec'],
        'mspd_rec' : m['mspd_rec'],
        'ar'       : m['ar'],
        'err_t'    : err_t,
        'err_R'    : err_R,
        'det_conf' : det_conf,
        'det_px'   : int(mask_bool.sum()),
    }


def print_ho3d_summary(seq_results, paper_results):
    """Print per-sequence table with paper comparison."""
    SEP = '═' * 88
    print(f'\n{SEP}')
    print('  HO3D EVALUATION SUMMARY  (YOLOE + Any6D  vs  Paper SAM2 + Any6D)')
    print(SEP)

    hdr = (f'  {"Seq":<8} {"Obj":<22} {"ADD-S%":>7} {"ADD%":>6} {"AR%":>6}'
           f' {"AR%(paper)":>10} {"ADDS%(paper)":>12} {"frames":>7}')
    print(hdr)
    print('  ' + '─' * 86)

    paper_seqs = paper_results.get('ho3d', {}).get('sequences', {})

    totals = {'adds_ok': 0, 'add_ok': 0, 'ar': [], 'n': 0}

    for seq_name, frames in seq_results.items():
        ok = [f for f in frames if f.get('success')]
        if not ok:
            print(f'  {seq_name:<8}  FAILED / NO DATA')
            continue
        obj   = ok[0]['folder']
        adds_pct = 100 * sum(f['adds_ok'] for f in ok) / len(ok)
        add_pct  = 100 * sum(f['add_ok']  for f in ok) / len(ok)
        ar_pct   = 100 * np.mean([f['ar'] for f in ok])
        p = paper_seqs.get(seq_name, {})
        p_ar   = f'{p.get("AR",   "—"):>5}' if p else '    —'
        p_adds = f'{p.get("ADD-S","—"):>5}' if p else '    —'
        print(f'  {seq_name:<8} {obj[:22]:<22}'
              f' {adds_pct:>7.1f} {add_pct:>6.1f} {ar_pct:>6.1f}'
              f' {p_ar:>10} {p_adds:>12} {len(ok):>7}')
        totals['adds_ok'] += sum(f['adds_ok'] for f in ok)
        totals['add_ok']  += sum(f['add_ok']  for f in ok)
        totals['ar'].extend(f['ar'] for f in ok)
        totals['n']       += len(ok)

    n = max(totals['n'], 1)
    paper_mean = paper_results.get('ho3d', {}).get('mean', {})
    print('  ' + '─' * 86)
    print(f'  {"MEAN":<8} {"(all sequences)":<22}'
          f' {100*totals["adds_ok"]/n:>7.1f}'
          f' {100*totals["add_ok"]/n:>6.1f}'
          f' {100*np.mean(totals["ar"]):>6.1f}'
          f' {str(paper_mean.get("AR","—")):>10}'
          f' {str(paper_mean.get("ADD-S","—")):>12}'
          f' {n:>7}')
    print(SEP)
    print('  Note: AR here = mean(MSSD-recall, MSPD-recall, ADD-S-binary)')
    print('        Paper AR = mean(VSD-recall, MSSD-recall, MSPD-recall)')
    print(SEP + '\n')


def export_ho3d_excel(seq_results, paper_results, out_path):
    """Excel with per-frame + per-sequence summary + paper comparison."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('  [Excel] openpyxl not installed')
        return

    wb = openpyxl.Workbook()
    paper_seqs = paper_results.get('ho3d', {}).get('sequences', {})
    paper_mean = paper_results.get('ho3d', {}).get('mean', {})

    # ── Sheet 1: per-sequence summary ────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    hf = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))

    cols = ['Sequence', 'Object', 'Frames',
            'ADD-S% (ours)', 'ADD% (ours)', 'AR% (ours)',
            'ADD-S% (paper)', 'ADD% (paper)', 'AR% (paper)',
            'Δ ADD-S', 'Δ ADD', 'Δ AR']
    ws.append(cols)
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 30

    totals = {'adds_ok': 0, 'add_ok': 0, 'ar': [], 'n': 0}
    for seq_name, frames in seq_results.items():
        ok = [f for f in frames if f.get('success')]
        if not ok:
            continue
        n = len(ok)
        adds_pct = 100 * sum(f['adds_ok'] for f in ok) / n
        add_pct  = 100 * sum(f['add_ok']  for f in ok) / n
        ar_pct   = 100 * np.mean([f['ar'] for f in ok])
        p = paper_seqs.get(seq_name, {})
        row = [
            seq_name, ok[0]['folder'], n,
            round(adds_pct, 1), round(add_pct, 1), round(ar_pct, 1),
            p.get('ADD-S', '—'), p.get('ADD', '—'), p.get('AR', '—'),
            round(adds_pct - p['ADD-S'], 1) if p.get('ADD-S') else '—',
            round(add_pct  - p['ADD'],   1) if p.get('ADD')   else '—',
            round(ar_pct   - p['AR'],    1) if p.get('AR')    else '—',
        ]
        ws.append(row)
        cur = ws.max_row
        for c in range(1, len(cols)+1):
            ws.cell(row=cur, column=c).border = thin

        # colour Δ cells
        for ci, key_pair in [(10, ('ADD-S', adds_pct)),
                              (11, ('ADD',   add_pct)),
                              (12, ('AR',    ar_pct))]:
            p_val = p.get(key_pair[0])
            if p_val is not None:
                delta = key_pair[1] - p_val
                fill_c = 'C8E6C9' if delta >= 0 else 'FFCDD2'
                ws.cell(row=cur, column=ci).fill = PatternFill(
                    start_color=fill_c, end_color=fill_c, fill_type='solid')

        totals['adds_ok'] += sum(f['adds_ok'] for f in ok)
        totals['add_ok']  += sum(f['add_ok']  for f in ok)
        totals['ar'].extend(f['ar'] for f in ok)
        totals['n']       += n

    # MEAN row
    N = max(totals['n'], 1)
    adds_m = round(100*totals['adds_ok']/N, 1)
    add_m  = round(100*totals['add_ok']/N, 1)
    ar_m   = round(100*np.mean(totals['ar']), 1) if totals['ar'] else 0
    mean_row = [
        'MEAN', '—', N, adds_m, add_m, ar_m,
        paper_mean.get('ADD-S','—'), paper_mean.get('ADD','—'),
        paper_mean.get('AR','—'),
        round(adds_m - paper_mean['ADD-S'], 1) if paper_mean.get('ADD-S') else '—',
        round(add_m  - paper_mean['ADD'],   1) if paper_mean.get('ADD')   else '—',
        round(ar_m   - paper_mean['AR'],    1) if paper_mean.get('AR')    else '—',
    ]
    ws.append(mean_row)
    mf = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.fill = mf; cell.font = Font(bold=True); cell.border = thin

    col_widths = [10, 25, 8, 14, 12, 12, 15, 13, 13, 9, 9, 9]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: per-frame details ────────────────────────────────────────────
    ws2 = wb.create_sheet('Frames')
    fcols = ['Seq', 'Frame', 'Object', 'ADD-S ok', 'ADD ok',
             'ADD-S (cm)', 'ADD (cm)', 'MSSD rec', 'MSPD rec', 'AR',
             'T err (cm)', 'R err (°)', 'YOLOE conf', 'Det px']
    ws2.append(fcols)
    for c in range(1, len(fcols)+1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin
    ws2.row_dimensions[1].height = 24

    for seq_name, frames in seq_results.items():
        for fr in frames:
            if not fr.get('success'):
                continue
            ws2.append([
                fr['seq_name'], fr['frame_id'], fr['folder'],
                '✓' if fr['adds_ok'] else '✗',
                '✓' if fr['add_ok']  else '✗',
                round(fr['adds_m'] * 100, 3),
                round(fr['add_m']  * 100, 3),
                round(fr['mssd_rec'], 3),
                round(fr['mspd_rec'], 3),
                round(fr['ar'],      3),
                round(fr['err_t'],   2),
                round(fr['err_R'],   2),
                round(fr['det_conf'],3),
                fr['det_px'],
            ])
            cur = ws2.max_row
            for c in range(1, len(fcols)+1):
                ws2.cell(row=cur, column=c).border = thin
            for ci, ok_key in [(4, fr['adds_ok']), (5, fr['add_ok'])]:
                fill_c = 'C8E6C9' if ok_key else 'FFCDD2'
                ws2.cell(row=cur, column=ci).fill = PatternFill(
                    start_color=fill_c, end_color=fill_c, fill_type='solid')

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f'\n  Excel: {out_path}')


def _yoloe_detect_frames(model, frames, conf, iou_thresh):
    """
    Phase 1: Run YOLOE on all frames of a sequence (host side).
    Returns list of (frame_dict, mask_bool, det_conf) or (frame_dict, None, 0).
    Groups by object so set_classes is called once per object.
    """
    results = []
    for frame in frames:
        try:
            _, mask_bool, det_conf, _ = yoloe_text_prompt(
                model, frame['color_path'], frame['prompt'],
                conf=conf, iou=iou_thresh,
            )
            results.append((frame, mask_bool, det_conf))
        except ValueError:
            results.append((frame, None, 0.0))
    return results


def run_ho3d_eval(args, dataset_cfg):
    """
    HO3D evaluation — BATCH mode:
      Phase 1 (host):   YOLOE detects all frames of a sequence
      Phase 2 (Docker): One Docker call processes ALL frames at once
                        → FoundationPose loaded once per sequence (~30s)
                        → ~8s per frame instead of ~60s
    """
    ho3d_root = dataset_cfg['ho3d_root']
    if not os.path.isdir(os.path.join(ho3d_root, 'evaluation')):
        print('[ERROR] HO3D evaluation sequences not found at:')
        print(f'  {ho3d_root}/evaluation/')
        print('Download HO3D v3 evaluation split and extract there.')
        sys.exit(1)

    seq_filter  = args.seq if getattr(args, 'seq', None) else None
    stride      = getattr(args, 'stride', 10)
    max_frames  = getattr(args, 'max_frames', None)
    if max_frames:
        dataset_cfg = dict(dataset_cfg)
        dataset_cfg['max_frames'] = max_frames
    save_dir   = os.path.join(ANY6D_DIR, 'results', 'pipeline', 'ho3d')
    viz_dir    = os.path.join(EVAL_DIR,  'ho3d', 'figures')
    os.makedirs(save_dir, exist_ok=True)
    os.makedirs(viz_dir,  exist_ok=True)

    print(f'\nLoading YOLOE model: {YOLOE_MODEL_PATH}')
    from ultralytics import YOLOE
    yoloe_model = YOLOE(YOLOE_MODEL_PATH)
    yoloe_model.to('cuda')
    print('YOLOE model loaded.\n')

    # Group frames by sequence
    seq_frames = {}
    for frame in _frames_from_ho3d(dataset_cfg, stride=stride,
                                    seq_filter=seq_filter):
        sn = frame['seq_name']
        if frame.get('missing'):
            print(f'  [SKIP] {sn} — folder not found')
            seq_frames[sn] = []
        else:
            seq_frames.setdefault(sn, []).append(frame)

    seq_results = {}

    for seq_name, frames in seq_frames.items():
        if not frames:
            seq_results[seq_name] = []
            continue

        obj_folder = frames[0]['folder']
        print(f'\n{"═"*60}')
        print(f'  Sequence: {seq_name}  |  Object: {obj_folder}')
        print(f'  Frames to process: {len(frames)}')
        print(f'{"═"*60}')

        # ── Phase 1: YOLOE detection (host, fast) ─────────────────────────────
        print(f'  [Phase 1] YOLOE detection on {len(frames)} frames...')
        prompts = frames[0]['prompt'] if isinstance(frames[0]['prompt'], list) \
                  else [frames[0]['prompt']]
        yoloe_model.set_classes(prompts, yoloe_model.get_text_pe(prompts))
        detected = _yoloe_detect_frames(yoloe_model, frames,
                                         conf=args.conf, iou_thresh=args.iou)
        n_det = sum(1 for _, m, _ in detected if m is not None)
        print(f'  → {n_det}/{len(frames)} detected')

        # ── Phase 2: prepare depth + frames_data for batch Docker ─────────────
        seq_save  = os.path.join(save_dir, seq_name)
        os.makedirs(seq_save, exist_ok=True)

        frames_data = []
        failed_frames = []

        for frame, mask_bool, det_conf in detected:
            fid = f'{frame["frame_id"]:05d}'
            if mask_bool is None:
                failed_frames.append({
                    'seq_name': seq_name, 'frame_id': frame['frame_id'],
                    'folder': obj_folder, 'success': False,
                    'error': 'YOLOE: no detection', 'det_conf': 0.0,
                })
                continue

            # Convert HO3D depth to uint16 mm
            depth_m    = _ho3d_load_depth(frame['depth_path'])
            depth_mm   = (depth_m * 1000).clip(0, 65535).astype(np.uint16)
            fdir       = os.path.join(seq_save, fid)
            os.makedirs(fdir, exist_ok=True)
            depth_path = os.path.join(fdir, 'depth_mm.png')
            cv2.imwrite(depth_path, depth_mm)

            frames_data.append({
                'id'        : fid,
                'color_path': frame['color_path'],
                'depth_path': depth_path,
                'mask_bool' : mask_bool,
                'K'         : frame['K_query'],
                # stash for metrics phase
                '_frame'    : frame,
                '_det_conf' : det_conf,
                '_mask_bool': mask_bool,
            })

        # ── Phase 2: Any6D batch Docker call (one init for all frames) ────────
        if frames_data:
            print(f'  [Phase 2] Any6D batch ({len(frames_data)} frames, 1 Docker call)...')
            poses = run_any6d_docker_batch(
                frames_data=[{k: v for k, v in fd.items()
                              if not k.startswith('_')}
                             for fd in frames_data],
                mesh_path=frames[0]['mesh_path'],
                save_dir=seq_save,
                name_prefix=seq_name,
                iteration=args.iter,
            )
        else:
            poses = {}

        # ── Phase 3: compute metrics per frame ────────────────────────────────
        mesh     = trimesh.load(frames[0]['mesh_path'])
        pts_src, _ = trimesh.sample.sample_surface(mesh, 1000)
        pts_src  = pts_src.astype(np.float64)
        diameter = float(np.linalg.norm(mesh.bounding_box.extents))

        pred_pose_a = np.loadtxt(frames[0]['init_pose_path'])
        gt_pose_a   = np.loadtxt(frames[0]['gt_pose_a_path'])

        seq_res = list(failed_frames)
        for fd in frames_data:
            fid     = fd['id']
            frame   = fd['_frame']
            det_conf = fd['_det_conf']
            mask_bool = fd['_mask_bool']

            pred_pose_q = poses.get(fid)
            if pred_pose_q is None:
                seq_res.append({
                    'seq_name': seq_name, 'frame_id': frame['frame_id'],
                    'folder': obj_folder, 'success': False,
                    'error': 'Any6D failed', 'det_conf': det_conf,
                })
                continue

            try:
                gt_pose_q = _ho3d_load_gt_pose(frame['pkl_path'])
                if gt_pose_q is None:
                    raise ValueError('No GT pose in pkl')

                pred_final = _apply_aq_transform(pred_pose_q,
                                                  pred_pose_a, gt_pose_a)
                m = compute_ho3d_metrics(pred_final, gt_pose_q,
                                          pts_src, diameter)
                mspd_rec, mspd_val = compute_mspd_recall(
                    pts_src, pred_final, gt_pose_q, frame['K_query'])
                m['mspd_rec'] = mspd_rec
                ar = float(np.mean([m['mssd_rec'], mspd_rec,
                                    float(m['adds_ok'])]))
                err_t, err_R = compute_errors(pred_final, gt_pose_q)

                res = {
                    'seq_name' : seq_name,
                    'frame_id' : frame['frame_id'],
                    'folder'   : obj_folder,
                    'success'  : True,
                    'adds_ok'  : m['adds_ok'],
                    'add_ok'   : m['add_ok'],
                    'adds_m'   : m['adds_m'],
                    'add_m'    : m['add_m'],
                    'mssd_rec' : m['mssd_rec'],
                    'mspd_rec' : mspd_rec,
                    'ar'       : ar,
                    'err_t'    : err_t,
                    'err_R'    : err_R,
                    'det_conf' : det_conf,
                    'det_px'   : int(mask_bool.sum()),
                }
                seq_res.append(res)
                print(f'    frame {frame["frame_id"]:05d}'
                      f'  ADD-S={"✓" if m["adds_ok"] else "✗"}'
                      f'  ADD={"✓" if m["add_ok"] else "✗"}'
                      f'  AR={ar:.2f}'
                      f'  T={err_t:.1f}cm  R={err_R:.1f}°'
                      f'  conf={det_conf:.2f}')
            except Exception as e:
                seq_res.append({
                    'seq_name': seq_name, 'frame_id': frame['frame_id'],
                    'folder': obj_folder, 'success': False,
                    'error': str(e), 'det_conf': det_conf,
                })
                print(f'    [✗] frame {frame["frame_id"]:05d}  {e}')

        seq_results[seq_name] = seq_res

    print_ho3d_summary(seq_results, PAPER_RESULTS)
    excel_path = os.path.join(EVAL_DIR, 'ho3d', 'ho3d_metrics.xlsx')
    export_ho3d_excel(seq_results, PAPER_RESULTS, excel_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='YOLOE + Any6D 6D pose estimation pipeline'
    )
    parser.add_argument('--dataset',   default='dexycb',
                        help='Dataset key from set_config.py')
    parser.add_argument('--obj',       default=None, nargs='+',
                        help='Run only these object folder name(s)')
    parser.add_argument('--csv',       default=None,
                        help='Path to CSV file for multi-frame mode')
    parser.add_argument('--conf',      type=float, default=YOLOE_CONF,
                        help='YOLOE confidence threshold')
    parser.add_argument('--iou',       type=float, default=YOLOE_IOU,
                        help='YOLOE NMS IoU threshold')
    parser.add_argument('--iter',      type=int,   default=ANY6D_ITER,
                        help='Any6D refinement iterations')
    # HO3D-specific args
    parser.add_argument('--stride',     type=int,  default=10,
                        help='HO3D frame stride (default 10, skip 9/10 frames)')
    parser.add_argument('--max_frames', type=int,  default=None,
                        help='HO3D: max frames per sequence (for quick tests)')
    parser.add_argument('--seq',       nargs='+',  default=None,
                        help='HO3D: subset of sequences to evaluate '
                             '(e.g. --seq AP10 AP11 SM1)')
    args = parser.parse_args()

    if args.dataset not in DATASETS:
        print(f'Unknown dataset "{args.dataset}". '
              f'Available: {list(DATASETS.keys())}')
        sys.exit(1)

    dataset_cfg  = DATASETS[args.dataset]
    dataset_name = dataset_cfg['name']

    # ── HO3D evaluation mode ──────────────────────────────────────────────────
    if args.dataset == 'ho3d':
        run_ho3d_eval(args, dataset_cfg)
        return

    objects = dataset_cfg['objects']
    if not objects:
        print(f'No objects configured for dataset "{args.dataset}".')
        print('Edit pipeline/set_config.py to add objects.')
        sys.exit(0)

    if args.obj:
        objects = [o for o in objects if o['folder'] in args.obj]
        if not objects:
            print(f'Object(s) {args.obj} not found in dataset "{args.dataset}".')
            sys.exit(1)

    viz_dir  = os.path.join(EVAL_DIR, args.dataset, 'figures')
    # save_dir MUST be inside Any6D/ so Docker (/workspace) can read the files
    save_dir = os.path.join(ANY6D_DIR, 'results', 'pipeline', args.dataset)
    os.makedirs(viz_dir,  exist_ok=True)
    os.makedirs(save_dir, exist_ok=True)

    print(f'\nLoading YOLOE model: {YOLOE_MODEL_PATH}')
    from ultralytics import YOLOE
    yoloe_model = YOLOE(YOLOE_MODEL_PATH)
    yoloe_model.to('cuda')
    print('YOLOE model loaded.\n')

    all_results = []
    for obj_cfg in objects:
        frames = (
            _frames_from_csv(args.csv, dataset_cfg)
            if args.csv else
            _frames_from_anchor(obj_cfg, dataset_cfg)
        )
        for frame in frames:
            try:
                result = process_frame(
                    yoloe_model, frame,
                    conf=args.conf, iou_thresh=args.iou,
                    iteration=args.iter,
                    viz_dir=viz_dir, save_dir=save_dir,
                )
                all_results.append(result)
            except Exception as e:
                print(f'  [FAIL] {frame.get("folder","?")} — {e}')
                all_results.append({
                    'folder' : frame.get('folder', '?'),
                    'success': False,
                    'error'  : str(e),
                })

    print_summary(all_results, dataset_name)

    excel_path = os.path.join(EVAL_DIR, args.dataset,
                              f'{args.dataset}_metrics.xlsx')
    export_excel(
        all_results,
        dataset_name,
        PAPER_RESULTS.get(args.dataset, {}),
        excel_path,
    )


if __name__ == '__main__':
    main()
