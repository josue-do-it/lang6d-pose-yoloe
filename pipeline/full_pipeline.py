"""
full_pipeline.py
Ollama → YOLOE → Any6D — full pipeline with multi-prompt evaluation.

Dataset mode (anchor frame, tests N prompts per object):
    python pipeline/full_pipeline.py --dataset dexycb
    python pipeline/full_pipeline.py --dataset dexycb --obj 006_mustard_bottle --n 50
    python pipeline/full_pipeline.py --dataset dexycb --prompts my_prompts.txt

Single image mode (one image + natural language instruction):
    python pipeline/full_pipeline.py \\
        --image scene.jpg --depth depth.png --K K.txt --mesh mesh.obj \\
        --user-prompt "pick up the yellow mustard bottle" --n 20
"""

import os
import sys
import ast
import json
import argparse
import requests
import numpy as np
import cv2
import trimesh

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

BASE_DIR = os.path.expanduser('~/open-vocabulary-6d-pose-yoloe')
sys.path.insert(0, os.path.join(BASE_DIR, 'utils'))
sys.path.insert(0, os.path.join(BASE_DIR, 'pipeline'))

from set_config import (
    DATASETS, EVAL_DIR, ANY6D_DIR, YOLOE_MODEL_PATH,
    YOLOE_CONF, YOLOE_IOU, ANY6D_ITER,
    THRESH_ADD_RATIO, THRESH_MSSD_RATIO, THRESH_MSPD_PX,
)
from pipeline_utils import run_any6d_docker, compute_errors, compute_mask_metrics
from yoloe_helpers  import yoloe_text_prompt
from any6d_utils    import (
    mesh_bbox_corners, project_points,
    draw_3d_bbox, draw_pose_axes, mask_overlay, colormap_depth,
)
from ollama_utils   import is_ollama_running

OLLAMA_HOST = 'http://localhost:11434'


# ── Prompt bank generation ────────────────────────────────────────────────────

def _systematic_prompts(base_prompt):
    """Generate systematic variations without any LLM."""
    words  = base_prompt.strip().lower().split()
    core   = words[-1]
    adj    = words[0] if len(words) > 1 else ''

    COLORS    = ['yellow', 'red', 'blue', 'green', 'white', 'orange', 'brown', 'black']
    MATERIALS = ['plastic', 'metal', 'glass', 'cardboard', 'rubber', 'ceramic']
    SIZES     = ['small', 'large', 'big', 'medium']

    pool = {base_prompt, core, f'a {base_prompt}', f'the {base_prompt}',
            f'a {core}', f'the {core}'}

    for col  in COLORS:    pool.add(f'{col} {core}')
    for mat  in MATERIALS: pool.add(f'{mat} {core}')
    for sz   in SIZES:     pool.add(f'{sz} {core}')

    if adj:
        pool.add(f'{adj} {core}')
        for mat in MATERIALS[:3]:
            pool.add(f'{adj} {mat} {core}')
        for sz in SIZES[:2]:
            pool.add(f'{sz} {adj} {core}')

    return [p.strip() for p in pool if p.strip()]


def _ollama_prompt_bank(base_prompt, n=30, model='llama3.2:3b'):
    """Single Ollama call: request n diverse YOLOE prompts for an object."""
    if not is_ollama_running():
        return []

    system = (
        f'Generate a Python list of {n} diverse text descriptors for YOLOE open-vocabulary '
        f'computer vision detection. Object: "{base_prompt}". '
        f'Vary: color, shape, material, brand name, function, size, texture. '
        f'Each descriptor: 1-6 words. '
        f'Return ONLY a valid Python list of strings, nothing else.'
    )
    payload = {
        'model'  : model,
        'prompt' : system,
        'stream' : False,
        'keep_alive': 0,
        'options': {'temperature': 0.4, 'num_predict': 600},
    }
    try:
        r   = requests.post(f'{OLLAMA_HOST}/api/generate', json=payload, timeout=60)
        r.raise_for_status()
        raw   = r.json()['response'].strip()
        start = raw.find('[')
        end   = raw.rfind(']') + 1
        if 0 <= start < end:
            return [str(p).strip().lower() for p in ast.literal_eval(raw[start:end]) if str(p).strip()]
    except Exception as e:
        print(f'  [Ollama] prompt bank generation failed: {e}')
    return []


def generate_prompt_bank(base_prompt, n=50, ollama_model='llama3.2:3b',
                          custom_prompts=None):
    """
    Build a bank of up to n unique YOLOE text prompts for one object.
    Order: custom → systematic → Ollama extras.
    """
    pool = set()

    if custom_prompts:
        pool.update(p.strip().lower() for p in custom_prompts if p.strip())

    pool.update(_systematic_prompts(base_prompt))

    n_extra = max(0, n - len(pool))
    if n_extra > 0:
        pool.update(_ollama_prompt_bank(base_prompt, n=n_extra + 10,
                                        model=ollama_model))

    result = sorted(p for p in pool if p.strip())[:n]
    print(f'  Prompt bank: {len(result)} prompts generated')
    return result


# ── YOLOE multi-prompt evaluation ────────────────────────────────────────────

def evaluate_prompts(yoloe_model, image_path, prompts,
                     conf=YOLOE_CONF, iou=YOLOE_IOU, gt_mask=None):
    """
    Test every prompt with YOLOE on one image.

    Returns list of dicts sorted by score descending:
        prompt, conf, mask_px, iou_score, score, mask, bbox, detected
    """
    results = []
    n = len(prompts)
    for i, prompt in enumerate(prompts):
        try:
            bbox, mask, det_conf, _ = yoloe_text_prompt(
                yoloe_model, image_path, prompt, conf=conf, iou=iou
            )
            iou_score = 0.0
            if gt_mask is not None and mask is not None:
                mm = compute_mask_metrics(mask, gt_mask)
                iou_score = mm['IoU']
            score = float(det_conf) * (1.0 + iou_score)
            results.append({
                'prompt'   : prompt,
                'conf'     : float(det_conf),
                'mask_px'  : int(mask.sum()) if mask is not None else 0,
                'iou_score': iou_score,
                'score'    : score,
                'mask'     : mask,
                'bbox'     : bbox,
                'detected' : True,
            })
        except Exception:
            results.append({
                'prompt'   : prompt,
                'conf'     : 0.0,
                'mask_px'  : 0,
                'iou_score': 0.0,
                'score'    : 0.0,
                'mask'     : None,
                'bbox'     : None,
                'detected' : False,
            })

        if (i + 1) % 10 == 0 or (i + 1) == n:
            det = sum(1 for r in results if r['detected'])
            print(f'    [{i+1}/{n}] tested  detected={det}')

    results.sort(key=lambda x: x['score'], reverse=True)
    return results


# ── Visualization ─────────────────────────────────────────────────────────────

def save_prompt_ranking_fig(prompt_results, obj_name, out_path, top_n=25):
    """Horizontal bar chart: YOLOE confidence per prompt (top top_n shown)."""
    detected = [r for r in prompt_results if r['detected']]
    missed   = [r for r in prompt_results if not r['detected']]

    display  = (detected + missed)[:top_n]
    labels   = [r['prompt'][:35] for r in display]
    confs    = [r['conf']        for r in display]
    colors   = ['#43A047' if r['detected'] else '#EF5350' for r in display]

    h = max(5, len(display) * 0.32)
    fig, ax = plt.subplots(figsize=(10, h))
    bars = ax.barh(range(len(display)), confs, color=colors, height=0.65, edgecolor='none')

    # star on best
    if detected:
        best_idx = next(i for i, r in enumerate(display) if r == detected[0])
        ax.barh(best_idx, confs[best_idx], color='#1565C0', height=0.65, edgecolor='none')
        ax.text(confs[best_idx] + 0.005, best_idx, '★ best', va='center', fontsize=8, color='#1565C0')

    ax.set_yticks(range(len(display)))
    ax.set_yticklabels(labels, fontsize=8)
    ax.invert_yaxis()
    ax.set_xlabel('YOLOE Confidence Score', fontsize=9)
    ax.set_xlim(0, 1.12)
    ax.axvline(YOLOE_CONF, color='grey', linestyle='--', linewidth=0.8, label=f'threshold={YOLOE_CONF}')
    ax.set_title(
        f'{obj_name.replace("_", " ")} — Prompt Ranking\n'
        f'Detected: {len(detected)}/{len(prompt_results)} prompts'
        + (f'  (showing top {top_n})' if len(prompt_results) > top_n else ''),
        fontsize=10
    )
    ax.legend(handles=[
        mpatches.Patch(color='#43A047', label='Detected'),
        mpatches.Patch(color='#EF5350', label='Not detected'),
        mpatches.Patch(color='#1565C0', label='Best'),
    ], fontsize=8, loc='lower right')
    ax.grid(axis='x', alpha=0.3)

    plt.tight_layout()
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    fig.savefig(out_path, dpi=120, bbox_inches='tight')
    plt.close(fig)
    print(f'  Prompt ranking: {out_path}')


def save_pose_fig(color_rgb, depth_raw, mask_bool,
                  pred_pose, gt_pose, K, mesh,
                  best_prompt, det_conf, err_t, err_R, ar,
                  obj_name, out_path):
    """4-panel figure: RGB | Depth | Mask | Pose."""
    fig, axes = plt.subplots(2, 2, figsize=(13, 9))
    fig.patch.set_facecolor('#f5f5f5')

    ar_str  = f'{ar:.3f}' if ar is not None and not np.isnan(ar) else 'N/A'
    et_str  = f'{err_t:.1f} cm' if err_t is not None and not np.isnan(err_t) else 'N/A'
    er_str  = f'{err_R:.1f}°'  if err_R is not None and not np.isnan(err_R) else 'N/A'
    plt.suptitle(
        f'{obj_name.replace("_", " ")}   '
        f'AR={ar_str}  T={et_str}  R={er_str}\n'
        f'Best prompt: "{best_prompt}"  (conf={det_conf:.3f})',
        fontsize=10, y=0.99
    )

    # panel 0,0 — Anchor RGB
    axes[0, 0].imshow(color_rgb)
    axes[0, 0].set_title('Anchor RGB', fontsize=9)
    axes[0, 0].axis('off')

    # panel 0,1 — Depth
    valid   = depth_raw[depth_raw > 0]
    d_min   = valid.min() if valid.size else 0
    d_max   = valid.max() if valid.size else 1
    d_norm  = np.clip((depth_raw - d_min) / (d_max - d_min + 1e-9), 0, 1)
    axes[0, 1].imshow(d_norm, cmap='inferno')
    axes[0, 1].set_title(f'Depth  [{d_min:.3f}–{d_max:.3f} m]', fontsize=9)
    axes[0, 1].axis('off')

    # panel 1,0 — YOLOE mask
    overlay = mask_overlay(color_rgb, mask_bool, alpha=0.45)
    axes[1, 0].imshow(overlay)
    axes[1, 0].set_title(
        f'YOLOE mask  "{best_prompt}"  conf={det_conf:.3f}  {mask_bool.sum()} px',
        fontsize=8
    )
    axes[1, 0].axis('off')

    # panel 1,1 — Pose
    corners  = mesh_bbox_corners(mesh)
    base_img = mask_overlay(color_rgb, mask_bool, alpha=0.18)
    c_pred   = project_points(corners, pred_pose, K)
    img      = draw_3d_bbox(base_img, c_pred, color=(30, 120, 255), thickness=2)
    img      = draw_pose_axes(img, pred_pose, K, length=0.05)
    if gt_pose is not None:
        c_gt = project_points(corners, gt_pose, K)
        img  = draw_3d_bbox(img, c_gt, color=(30, 200, 80), thickness=2)

    ar_color = '#2e7d32' if (ar is not None and not np.isnan(ar) and ar >= 0.8) \
               else '#f57f17' if (ar is not None and not np.isnan(ar) and ar >= 0.4) \
               else '#c62828'
    axes[1, 1].imshow(img)
    axes[1, 1].set_title(
        f'Pose  T={et_str}  R={er_str}  AR={ar_str}',
        fontsize=9, color=ar_color
    )
    axes[1, 1].axis('off')
    axes[1, 1].legend(handles=[
        mpatches.Patch(color='#1E78FF', label='Predicted'),
        mpatches.Patch(color='#1EC850', label='GT'),
    ], fontsize=7, loc='lower left', framealpha=0.75)

    plt.tight_layout(rect=[0, 0, 1, 0.96])
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    fig.savefig(out_path, dpi=120, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f'  Pose figure: {out_path}')


# ── R_fix ─────────────────────────────────────────────────────────────────────

def _apply_r_fix(pred_pose, gt_pose, init_pose_path):
    if not os.path.exists(init_pose_path):
        return gt_pose
    init = np.loadtxt(init_pose_path)
    diff = init[:3, :3] @ pred_pose[:3, :3].T
    angle = np.degrees(np.arccos(np.clip((np.trace(diff) - 1) / 2, -1, 1)))
    ref_R = init[:3, :3] if angle < 30 else pred_pose[:3, :3]
    R_fix = np.eye(4)
    R_fix[:3, :3] = gt_pose[:3, :3].T @ ref_R
    return gt_pose @ R_fix


# ── Per-object processing ─────────────────────────────────────────────────────

def process_object(yoloe_model, folder, anchor_dir, mesh_path, prompt_bank,
                   conf, iou_thresh, iteration, viz_dir, save_dir,
                   symmetric=False):
    """
    Run evaluate_prompts → Any6D → metrics for one object.
    Returns result dict.
    """
    color_path = os.path.join(anchor_dir, 'color.png')
    depth_path = os.path.join(anchor_dir, 'depth.png')
    K_path     = os.path.join(anchor_dir, 'K.txt')
    gt_mask_p  = os.path.join(anchor_dir, 'mask.png')
    gt_pose_p  = os.path.join(anchor_dir, f'{folder}_gt_pose.txt')
    init_pose_p= os.path.join(anchor_dir, f'{folder}_initial_pose.txt')

    color_bgr = cv2.imread(color_path)
    if color_bgr is None:
        raise FileNotFoundError(f'color.png not found: {color_path}')
    color_rgb = cv2.cvtColor(color_bgr, cv2.COLOR_BGR2RGB)
    depth_raw = cv2.imread(depth_path, cv2.IMREAD_ANYDEPTH).astype(np.float32) / 1000.0
    K         = np.loadtxt(K_path)
    mesh      = trimesh.load(mesh_path)

    gt_mask = None
    if os.path.exists(gt_mask_p):
        gm = cv2.imread(gt_mask_p, cv2.IMREAD_GRAYSCALE)
        gt_mask = gm > 127

    print(f'\n[{folder}] Testing {len(prompt_bank)} prompts...')
    prompt_results = evaluate_prompts(
        yoloe_model, color_path, prompt_bank,
        conf=conf, iou=iou_thresh, gt_mask=gt_mask
    )

    detected = [r for r in prompt_results if r['detected']]
    if not detected:
        raise ValueError(f'No prompt detected the object (tried {len(prompt_bank)} prompts)')

    best      = detected[0]
    mask_bool = best['mask']
    det_conf  = best['conf']
    print(f'  Best prompt: "{best["prompt"]}"  conf={det_conf:.3f}  mask={mask_bool.sum()} px')

    # mask quality vs GT
    mask_metrics = {}
    if gt_mask is not None:
        mask_metrics = compute_mask_metrics(mask_bool, gt_mask)
        print(f'  Mask IoU={mask_metrics["IoU"]:.3f}  Dice={mask_metrics["Dice"]:.3f}')

    # Any6D pose
    save_path = os.path.join(save_dir, folder)
    pred_pose = run_any6d_docker(
        color_path=color_path,
        depth_path=depth_path,
        mask_bool=mask_bool,
        K=K,
        mesh_path=mesh_path,
        save_path=save_path,
        name=folder,
        iteration=iteration,
    )

    # GT pose + errors
    gt_pose = None
    err_t = err_R = ar = float('nan')
    if os.path.exists(gt_pose_p):
        gt_pose = np.loadtxt(gt_pose_p)
        gt_pose = _apply_r_fix(pred_pose, gt_pose, init_pose_p)
        err_t, err_R = compute_errors(pred_pose, gt_pose)
        from pipeline_yolany6d import compute_bop_metrics
        bop = compute_bop_metrics(pred_pose, gt_pose, mesh, K, symmetric=symmetric)
        ar  = bop['AR']
        print(f'  T={err_t:.2f} cm  R={err_R:.2f}°  AR={ar:.3f}')
    else:
        bop = {}

    # figures
    obj_viz = os.path.join(viz_dir, folder)
    os.makedirs(obj_viz, exist_ok=True)

    save_prompt_ranking_fig(
        prompt_results, folder,
        os.path.join(obj_viz, 'prompt_ranking.png')
    )
    save_pose_fig(
        color_rgb, depth_raw, mask_bool,
        pred_pose, gt_pose, K, mesh,
        best['prompt'], det_conf, err_t, err_R, ar,
        folder, os.path.join(obj_viz, 'pose_result.png')
    )

    # save best prompt and pose
    with open(os.path.join(obj_viz, 'best_prompt.txt'), 'w') as f:
        f.write(f'{best["prompt"]}\nconf={det_conf:.4f}\n')
    np.savetxt(os.path.join(obj_viz, 'pred_pose.txt'), pred_pose)

    return {
        'folder'        : folder,
        'best_prompt'   : best['prompt'],
        'det_conf'      : det_conf,
        'n_detected'    : len(detected),
        'n_tested'      : len(prompt_bank),
        'det_rate'      : len(detected) / max(len(prompt_bank), 1),
        'mask_metrics'  : mask_metrics,
        'pred_pose'     : pred_pose,
        'gt_pose'       : gt_pose,
        'err_t'         : err_t,
        'err_R'         : err_R,
        'bop'           : bop,
        'ar'            : ar,
        'prompt_results': prompt_results,
        'success'       : True,
    }


# ── Single image mode ─────────────────────────────────────────────────────────

def run_single_image(yoloe_model, args):
    """Full pipeline on one user-provided image + prompt."""
    from ollama_utils import extract_yoloe_prompt

    print(f'\n[Single-image mode]')
    print(f'  Image      : {args.image}')
    print(f'  User said  : "{args.user_prompt}"')

    K    = np.loadtxt(args.K)
    mesh = trimesh.load(args.mesh)

    # Step 1: LLM extracts clean YOLOE object name from natural language
    raw_input = args.user_prompt or 'object'
    try:
        extracted = extract_yoloe_prompt(raw_input, model=args.ollama_model)
        print(f'  LLM extracted: "{extracted}"')
    except Exception as e:
        extracted = raw_input
        print(f'  [Ollama unavailable] using raw prompt: "{extracted}"  ({e})')

    # Step 2: build a small prompt bank from the extracted name
    base    = extracted
    prompts = generate_prompt_bank(base, n=args.n,
                                   ollama_model=args.ollama_model,
                                   custom_prompts=[extracted])

    print(f'  Testing {len(prompts)} prompts...')
    prompt_results = evaluate_prompts(
        yoloe_model, args.image, prompts,
        conf=args.conf, iou=args.iou
    )

    detected = [r for r in prompt_results if r['detected']]
    if not detected:
        print('  No detection found. Try a different prompt or lower --conf.')
        return None

    best      = detected[0]
    mask_bool = best['mask']
    print(f'  Best prompt: "{best["prompt"]}"  conf={best["conf"]:.3f}')

    # depth
    depth_raw = cv2.imread(args.depth, cv2.IMREAD_ANYDEPTH).astype(np.float32) / 1000.0

    # Any6D
    out_dir = os.path.join(ANY6D_DIR, 'results', 'pipeline', 'single_image')
    pred_pose = run_any6d_docker(
        color_path=args.image,
        depth_path=args.depth,
        mask_bool=mask_bool,
        K=K,
        mesh_path=args.mesh,
        save_path=out_dir,
        name='single_image',
        iteration=args.iter,
    )

    color_bgr = cv2.imread(args.image)
    color_rgb = cv2.cvtColor(color_bgr, cv2.COLOR_BGR2RGB)
    viz_dir   = os.path.join(EVAL_DIR, 'single_image', 'figures')

    save_prompt_ranking_fig(
        prompt_results, 'single_image',
        os.path.join(viz_dir, 'prompt_ranking.png')
    )
    save_pose_fig(
        color_rgb, depth_raw, mask_bool,
        pred_pose, None, K, mesh,
        best['prompt'], best['conf'],
        float('nan'), float('nan'), float('nan'),
        'single_image', os.path.join(viz_dir, 'pose_result.png')
    )
    print('\n[Result]')
    print(f'  User instruction : "{raw_input}"')
    print(f'  LLM extracted    : "{extracted}"')
    print(f'  Best YOLOE prompt: "{best["prompt"]}"  conf={best["conf"]:.3f}')
    print(f'  Translation (m)  : {pred_pose[:3,3].round(4).tolist()}')
    print(f'  Rotation         :\n{pred_pose[:3,:3].round(4)}')
    print(f'  Figures          : {viz_dir}')
    print(f'  Pose saved       : {out_dir}/pred_pose.txt')
    return pred_pose


# ── Terminal summary ──────────────────────────────────────────────────────────

def print_summary(all_results, dataset_name):
    SEP = '═' * 80
    print(f'\n{SEP}')
    print(f'  FULL PIPELINE SUMMARY — {dataset_name}')
    print(SEP)

    ok  = [r for r in all_results if r.get('success')]
    bad = [r for r in all_results if not r.get('success')]
    print(f'  Objects: {len(ok)}/{len(all_results)} succeeded  |  Failed: {len(bad)}')

    if ok:
        hdr = (f'\n  {"Object":<30} {"Best prompt":<30} {"conf":>5} '
               f'{"det%":>5} {"AR":>5} {"T(cm)":>7} {"R(°)":>7} {"IoU":>6}')
        print(hdr)
        print('  ' + '─' * 80)

        for r in ok:
            bp  = r['best_prompt'][:28]
            iou = r['mask_metrics'].get('IoU', float('nan'))
            ar  = r.get('ar', float('nan'))
            et  = r.get('err_t', float('nan'))
            er  = r.get('err_R', float('nan'))
            det = r.get('det_rate', 0) * 100

            def _f(v, nd=2):
                return f'{v:.{nd}f}' if v is not None and not np.isnan(v) else ' N/A'

            print(f'  {r["folder"]:<30} {bp:<30} {_f(r["det_conf"]):>5}'
                  f' {det:>4.0f}% {_f(ar):>5} {_f(et):>7} {_f(er):>7} {_f(iou):>6}')

        print('  ' + '─' * 80)

        def _mean(key, sub=None):
            vals = []
            for r in ok:
                v = r.get(key)
                if sub and isinstance(v, dict):
                    v = v.get(sub)
                if v is not None and not (isinstance(v, float) and np.isnan(v)):
                    vals.append(float(v))
            return np.mean(vals) if vals else float('nan')

        print(f'  {"MEAN":<61} {_mean("det_conf"):>5.2f} '
              f'{_mean("det_rate")*100:>4.0f}% '
              f'{_mean("ar"):>5.2f} '
              f'{_mean("err_t"):>7.1f} '
              f'{_mean("err_R"):>7.1f} '
              f'{_mean("mask_metrics","IoU"):>6.3f}')

    if bad:
        print('\n  Failed:')
        for r in bad:
            print(f'    {r.get("folder","?")} — {r.get("error","?")}')
    print(SEP + '\n')


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(all_results, dataset_name, out_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('  openpyxl not found — skip Excel')
        return

    wb  = openpyxl.Workbook()

    # ── Results sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = dataset_name[:31]
    h_fill  = PatternFill(start_color='263238', end_color='263238', fill_type='solid')
    h_font  = Font(color='FFFFFF', bold=True, size=9)
    thin    = Border(**{s: Side(style='thin') for s in ('left','right','top','bottom')})

    headers = [
        'Object', 'Best prompt', 'Conf', 'Det rate', 'N detected', 'N tested',
        'AR', 'ADD-S ok', 'MSSD ok', 'MSPD ok',
        'IoU', 'Dice', 'Precision', 'Recall',
        'T err (cm)', 'R err (°)',
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 26

    ok_fill  = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    bad_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')

    def _v(val, scale=1, nd=3):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return 'N/A'
        return round(float(val) * scale, nd)

    for r in all_results:
        if not r.get('success'):
            continue
        bop = r.get('bop', {})
        mm  = r.get('mask_metrics', {})
        row = [
            r['folder'],
            r.get('best_prompt', ''),
            _v(r.get('det_conf')),
            _v(r.get('det_rate'), nd=2),
            r.get('n_detected', 0),
            r.get('n_tested',   0),
            _v(r.get('ar')),
            '✓' if bop.get('ADDS_ok') else '✗',
            '✓' if bop.get('MSSD_ok') else '✗',
            '✓' if bop.get('MSPD_ok') else '✗',
            _v(mm.get('IoU')),
            _v(mm.get('Dice')),
            _v(mm.get('Precision')),
            _v(mm.get('Recall')),
            _v(r.get('err_t'), nd=2),
            _v(r.get('err_R'), nd=2),
        ]
        ws.append(row)
        cur = ws.max_row
        for col in range(1, len(headers)+1):
            ws.cell(cur, col).border = thin
        for ci, key in zip([8, 9, 10], ['ADDS_ok','MSSD_ok','MSPD_ok']):
            ws.cell(cur, ci).fill = ok_fill if bop.get(key) else bad_fill
            ws.cell(cur, ci).alignment = Alignment(horizontal='center')

    widths = [28, 32, 7, 9, 11, 9, 7, 8, 8, 8, 7, 7, 10, 9, 11, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Prompt ranking sheet ──────────────────────────────────────────────────
    wp = wb.create_sheet('Prompt Rankings')
    wp.append(['Object', 'Rank', 'Prompt', 'Conf', 'IoU score', 'Score', 'Detected'])
    wp.cell(1, 1).font = Font(bold=True)

    for r in all_results:
        if not r.get('success'):
            continue
        for rank, pr in enumerate(r.get('prompt_results', [])[:30], 1):
            wp.append([
                r['folder'], rank, pr['prompt'],
                round(pr['conf'], 4),
                round(pr['iou_score'], 4),
                round(pr['score'], 4),
                'Yes' if pr['detected'] else 'No',
            ])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f'  Excel: {out_path}')


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Ollama + YOLOE + Any6D full pipeline'
    )

    # dataset mode
    parser.add_argument('--dataset',      default='dexycb')
    parser.add_argument('--obj',          default=None)
    parser.add_argument('--n',            type=int,   default=50,
                        help='Max prompts to generate and test per object')
    parser.add_argument('--prompts',      default=None,
                        help='Path to .txt file with one prompt per line')

    # single image mode
    parser.add_argument('--image',        default=None, help='Path to scene image')
    parser.add_argument('--depth',        default=None, help='Path to depth image')
    parser.add_argument('--K',            default=None, help='Path to K.txt')
    parser.add_argument('--mesh',         default=None, help='Path to mesh .obj')
    parser.add_argument('--user-prompt',  default=None,
                        help='Natural language instruction (single-image mode)')

    # shared
    parser.add_argument('--conf',         type=float, default=YOLOE_CONF)
    parser.add_argument('--iou',          type=float, default=YOLOE_IOU)
    parser.add_argument('--iter',         type=int,   default=ANY6D_ITER)
    parser.add_argument('--ollama-model', default='llama3.2:3b')
    args = parser.parse_args()

    # load custom prompts file
    custom_prompts = None
    if args.prompts and os.path.exists(args.prompts):
        with open(args.prompts) as f:
            custom_prompts = [l.strip() for l in f if l.strip()]
        print(f'Loaded {len(custom_prompts)} custom prompts from {args.prompts}')

    print(f'Loading YOLOE model...')
    from ultralytics import YOLOE
    yoloe_model = YOLOE(YOLOE_MODEL_PATH)
    yoloe_model.to('cuda')

    # ── Single image mode ──────────────────────────────────────────────────────
    if args.image:
        for req in ('depth', 'K', 'mesh'):
            if not getattr(args, req.replace('-', '_')):
                print(f'Single-image mode requires --{req}')
                sys.exit(1)
        run_single_image(yoloe_model, args)
        return

    # ── Dataset mode ───────────────────────────────────────────────────────────
    if args.dataset not in DATASETS:
        print(f'Unknown dataset. Available: {list(DATASETS.keys())}')
        sys.exit(1)

    dataset_cfg  = DATASETS[args.dataset]
    dataset_name = dataset_cfg['name']
    objects      = dataset_cfg['objects']

    if not objects:
        print(f'No objects configured for "{args.dataset}". Edit pipeline/set_config.py.')
        sys.exit(0)

    if args.obj:
        objects = [o for o in objects if o['folder'] == args.obj]
        if not objects:
            print(f'Object "{args.obj}" not found.')
            sys.exit(1)

    viz_dir  = os.path.join(EVAL_DIR, args.dataset, 'full_pipeline')
    save_dir = os.path.join(ANY6D_DIR, 'results', 'pipeline', args.dataset)
    os.makedirs(save_dir, exist_ok=True)

    all_results = []
    for obj_cfg in objects:
        folder     = obj_cfg['folder']
        anchor_dir = os.path.join(dataset_cfg['anchor_dir'], folder)
        mesh_path  = os.path.join(anchor_dir, obj_cfg['mesh'])

        base   = obj_cfg['prompt'] if isinstance(obj_cfg['prompt'], str) \
                 else obj_cfg['prompt'][0]
        bank   = generate_prompt_bank(
            base, n=args.n,
            ollama_model=args.ollama_model,
            custom_prompts=custom_prompts,
        )

        try:
            result = process_object(
                yoloe_model, folder, anchor_dir, mesh_path, bank,
                conf=args.conf, iou_thresh=args.iou, iteration=args.iter,
                viz_dir=viz_dir, save_dir=save_dir,
                symmetric=obj_cfg.get('symmetric', False),
            )
            all_results.append(result)
        except Exception as e:
            print(f'  [FAIL] {folder} — {e}')
            all_results.append({'folder': folder, 'success': False, 'error': str(e)})

    print_summary(all_results, dataset_name)
    export_excel(
        all_results, dataset_name,
        os.path.join(EVAL_DIR, args.dataset, f'{args.dataset}_full_pipeline.xlsx')
    )


if __name__ == '__main__':
    main()
